# app/services/export_generator.py
from typing import Dict, Any
import os
import csv
from datetime import datetime
from app.schemas.report import FinancialReportSummary, ExportOptions
import logging

logger = logging.getLogger(__name__)


class PDFReportGenerator:
    @staticmethod
    async def generate(
            report_data: FinancialReportSummary,
            options: ExportOptions,
            export_dir: str,
            export_id: str
    ) -> str:
        """Генерация PDF отчета"""
        try:
            from reportlab.lib.pagesizes import letter, A4
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
            from reportlab.lib import colors
            from reportlab.graphics.shapes import Drawing
            from reportlab.graphics.charts.piecharts import Pie
            from reportlab.graphics.charts.lineplots import LinePlot
            from reportlab.graphics.widgetbase import Widget

            logger.info(f"Generating PDF report with ID: {export_id}")

        except ImportError as e:
            logger.warning(f"ReportLab not available: {e}. Using text fallback.")
            return await PDFReportGenerator._generate_simple_text_report(
                report_data, options, export_dir, export_id
            )

        file_path = os.path.join(export_dir, f"financial_report_{export_id}.pdf")
        logger.info(f"PDF file path: {file_path}")

        try:
            # Создаем PDF документ
            doc = SimpleDocTemplate(
                file_path, 
                pagesize=A4,
                rightMargin=72,
                leftMargin=72,
                topMargin=72,
                bottomMargin=18
            )
            styles = getSampleStyleSheet()
            story = []

            # Заголовок
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=20,
                spaceAfter=30,
                alignment=1,  # Центрирование
                textColor=colors.darkblue
            )

            story.append(Paragraph(f"Financial Report", title_style))
            story.append(Paragraph(f"{report_data.period_name}", styles['Heading2']))
            story.append(Spacer(1, 20))

            # Основные показатели
            summary_data = [
                ['Metric', 'Amount'],
                ['Income', f"${report_data.income:,.2f}"],
                ['Expenses', f"${report_data.expenses:,.2f}"],
                ['Net Balance', f"${report_data.net_balance:,.2f}"],
                ['Savings', f"${report_data.savings:,.2f}"],
                ['Savings Rate', f"{report_data.savings_rate:.1f}%"],
                ['Transaction Count', f"{report_data.transaction_count:,}"],
            ]

            summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
            summary_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTSIZE', (0, 1), (-1, -1), 10),
            ]))

            story.append(Paragraph("Financial Summary", styles['Heading2']))
            story.append(summary_table)
            story.append(Spacer(1, 20))

            # Бюджет
            if options.include_budget_analysis and report_data.budget_total > 0:
                budget_data = [
                    ['Budget Metric', 'Value'],
                    ['Total Budget', f"${report_data.budget_total:,.2f}"],
                    ['Used', f"${report_data.budget_used:,.2f}"],
                    ['Remaining', f"${report_data.budget_remaining:,.2f}"],
                    ['Usage', f"{report_data.budget_percentage:.1f}%"],
                ]

                budget_table = Table(budget_data, colWidths=[3*inch, 2*inch])
                budget_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 12),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.lightblue),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('FONTSIZE', (0, 1), (-1, -1), 10),
                ]))

                story.append(Paragraph("Budget Analysis", styles['Heading2']))
                story.append(budget_table)
                story.append(Spacer(1, 20))

            # Категории расходов
            if options.include_categories_summary and report_data.spending_categories:
                story.append(Paragraph("Spending by Categories", styles['Heading2']))

                categories_data = [['Category', 'Amount', 'Percentage', 'Transactions']]
                for category in report_data.spending_categories[:15]:  # Топ 15
                    categories_data.append([
                        category.category_name[:20],  # Ограничиваем длину
                        f"${category.amount:,.2f}",
                        f"{category.percentage:.1f}%",
                        str(category.transaction_count)
                    ])

                categories_table = Table(categories_data, colWidths=[2.5*inch, 1.5*inch, 1*inch, 1*inch])
                categories_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.lightgreen),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('FONTSIZE', (0, 1), (-1, -1), 9),
                ]))

                story.append(categories_table)
                story.append(Spacer(1, 20))

            # Данные для графика доходы vs расходы (если включены графики)
            if options.include_charts and report_data.income_vs_expenses_chart and report_data.income_vs_expenses_chart.data_points:
                story.append(Paragraph("Income vs Expenses Trend", styles['Heading2']))
                
                # Создаем таблицу с данными графика
                chart_data = [['Date', 'Income', 'Expenses', 'Net']]
                for point in report_data.income_vs_expenses_chart.data_points[-10:]:  # Последние 10 точек
                    chart_data.append([
                        point.date.strftime('%m/%d'),
                        f"${point.income:,.0f}",
                        f"${point.expenses:,.0f}",
                        f"${point.net:,.0f}"
                    ])

                chart_table = Table(chart_data, colWidths=[1.5*inch, 1.5*inch, 1.5*inch, 1.5*inch])
                chart_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('FONTSIZE', (0, 1), (-1, -1), 9),
                ]))

                story.append(chart_table)
                story.append(Spacer(1, 20))

            # Недельный тренд
            if report_data.weekly_trend and report_data.weekly_trend.weeks:
                story.append(Paragraph("Weekly Spending Trend", styles['Heading2']))
                
                weekly_data = [['Week', 'Total Expenses']]
                for week in report_data.weekly_trend.weeks:
                    week_label = f"{week.week_start.strftime('%m/%d')} - {week.week_end.strftime('%m/%d')}"
                    weekly_data.append([
                        week_label,
                        f"${week.total_expenses:,.2f}"
                    ])

                # Добавляем среднее
                weekly_data.append([
                    "Average",
                    f"${report_data.weekly_trend.average_weekly_spending:,.2f}"
                ])

                weekly_table = Table(weekly_data, colWidths=[3*inch, 2*inch])
                weekly_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('FONTSIZE', (0, 1), (-1, -1), 9),
                    ('BACKGROUND', (-1, -1), (-1, -1), colors.yellow),  # Выделяем среднее
                ]))

                story.append(weekly_table)
                story.append(Spacer(1, 20))

            # Инсайты
            if options.include_insights and report_data.insights:
                story.append(Paragraph("Financial Insights", styles['Heading2']))

                for i, insight in enumerate(report_data.insights, 1):
                    insight_text = f"<b>{i}. {insight.title}:</b> {insight.message}"
                    story.append(Paragraph(insight_text, styles['Normal']))
                    story.append(Spacer(1, 8))

                story.append(Spacer(1, 12))

            # Дополнительная статистика
            if report_data.largest_expense:
                story.append(Paragraph("Additional Statistics", styles['Heading2']))
                
                stats_data = [
                    ['Statistic', 'Value'],
                    ['Average Transaction', f"${report_data.average_transaction_amount:,.2f}"],
                    ['Largest Expense', f"${report_data.largest_expense.get('amount', 0):,.2f}"],
                    ['Largest Expense Category', report_data.largest_expense.get('category', 'N/A')],
                    ['Most Spending Category', report_data.most_spending_category],
                ]

                stats_table = Table(stats_data, colWidths=[3*inch, 2*inch])
                stats_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('FONTSIZE', (0, 1), (-1, -1), 9),
                ]))

                story.append(stats_table)
                story.append(Spacer(1, 20))

            # Футер
            story.append(Spacer(1, 30))
            footer_text = f"Report generated on {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
            footer_style = ParagraphStyle(
                'Footer',
                parent=styles['Normal'],
                fontSize=10,
                textColor=colors.grey,
                alignment=1
            )
            story.append(Paragraph(footer_text, footer_style))

            # Сохраняем PDF
            logger.info(f"Building PDF document...")
            doc.build(story)
            
            # Проверяем что файл создался
            if os.path.exists(file_path):
                file_size = os.path.getsize(file_path)
                logger.info(f"PDF file created successfully. Size: {file_size} bytes")
                
                if file_size == 0:
                    logger.error("PDF file is empty!")
                    raise Exception("Generated PDF file is empty")
                    
                return file_path
            else:
                logger.error("PDF file was not created")
                raise Exception("PDF file was not created")

        except Exception as e:
            logger.error(f"Error creating PDF: {e}")
            # Fallback к текстовому файлу
            return await PDFReportGenerator._generate_simple_text_report(
                report_data, options, export_dir, export_id
            )

    @staticmethod
    async def _generate_simple_text_report(
            report_data: FinancialReportSummary,
            options: ExportOptions,
            export_dir: str,
            export_id: str
    ) -> str:
        """Генерация простого текстового отчета (если reportlab недоступен)"""
        file_path = os.path.join(export_dir, f"financial_report_{export_id}.txt")
        logger.info(f"Generating text report: {file_path}")

        try:
            with open(file_path, 'w', encoding='utf-8') as f:
                f.write(f"FINANCIAL REPORT - {report_data.period_name}\n")
                f.write("=" * 50 + "\n\n")

                f.write("SUMMARY:\n")
                f.write(f"Income: ${report_data.income:,.2f}\n")
                f.write(f"Expenses: ${report_data.expenses:,.2f}\n")
                f.write(f"Net Balance: ${report_data.net_balance:,.2f}\n")
                f.write(f"Savings: ${report_data.savings:,.2f}\n")
                f.write(f"Savings Rate: {report_data.savings_rate:.1f}%\n")
                f.write(f"Transaction Count: {report_data.transaction_count:,}\n\n")

                if options.include_budget_analysis and report_data.budget_total > 0:
                    f.write("BUDGET ANALYSIS:\n")
                    f.write(f"Total Budget: ${report_data.budget_total:,.2f}\n")
                    f.write(f"Used: ${report_data.budget_used:,.2f} ({report_data.budget_percentage:.1f}%)\n")
                    f.write(f"Remaining: ${report_data.budget_remaining:,.2f}\n\n")

                if options.include_categories_summary and report_data.spending_categories:
                    f.write("SPENDING BY CATEGORIES:\n")
                    for category in report_data.spending_categories:
                        f.write(f"{category.category_name}: ${category.amount:,.2f} ({category.percentage:.1f}%) - {category.transaction_count} transactions\n")
                    f.write("\n")

                if report_data.weekly_trend and report_data.weekly_trend.weeks:
                    f.write("WEEKLY SPENDING TREND:\n")
                    for week in report_data.weekly_trend.weeks:
                        f.write(f"{week.week_start.strftime('%m/%d')} - {week.week_end.strftime('%m/%d')}: ${week.total_expenses:,.2f}\n")
                    f.write(f"Average Weekly Spending: ${report_data.weekly_trend.average_weekly_spending:,.2f}\n")
                    f.write(f"Trend: {report_data.weekly_trend.trend_percentage:+.1f}%\n\n")

                if options.include_insights and report_data.insights:
                    f.write("INSIGHTS:\n")
                    for i, insight in enumerate(report_data.insights, 1):
                        f.write(f"{i}. {insight.title}: {insight.message}\n")
                    f.write("\n")

                if report_data.largest_expense:
                    f.write("ADDITIONAL STATISTICS:\n")
                    f.write(f"Average Transaction: ${report_data.average_transaction_amount:,.2f}\n")
                    f.write(f"Largest Expense: ${report_data.largest_expense.get('amount', 0):,.2f} ({report_data.largest_expense.get('category', 'N/A')})\n")
                    f.write(f"Most Spending Category: {report_data.most_spending_category}\n\n")

                f.write(f"Report generated on {datetime.now().strftime('%B %d, %Y at %I:%M %p')}\n")

            logger.info(f"Text report created successfully")
            return file_path

        except Exception as e:
            logger.error(f"Error creating text report: {e}")
            raise


class CSVReportGenerator:
    @staticmethod
    async def generate(
            report_data: FinancialReportSummary,
            options: ExportOptions,
            export_dir: str,
            export_id: str
    ) -> str:
        """Генерация CSV отчета"""
        file_path = os.path.join(export_dir, f"financial_report_{export_id}.csv")
        logger.info(f"Generating CSV report: {file_path}")

        try:
            with open(file_path, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.writer(csvfile)

                # Заголовок
                writer.writerow([f"Financial Report - {report_data.period_name}"])
                writer.writerow([f"Generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
                writer.writerow([])

                # Основные показатели
                writer.writerow(["FINANCIAL SUMMARY"])
                writer.writerow(["Metric", "Amount"])
                writer.writerow(["Income", f"{report_data.income:.2f}"])
                writer.writerow(["Expenses", f"{report_data.expenses:.2f}"])
                writer.writerow(["Net Balance", f"{report_data.net_balance:.2f}"])
                writer.writerow(["Savings", f"{report_data.savings:.2f}"])
                writer.writerow(["Savings Rate (%)", f"{report_data.savings_rate:.1f}"])
                writer.writerow(["Transaction Count", f"{report_data.transaction_count}"])
                writer.writerow([])

                # Бюджет
                if options.include_budget_analysis and report_data.budget_total > 0:
                    writer.writerow(["BUDGET ANALYSIS"])
                    writer.writerow(["Metric", "Amount"])
                    writer.writerow(["Total Budget", f"{report_data.budget_total:.2f}"])
                    writer.writerow(["Used", f"{report_data.budget_used:.2f}"])
                    writer.writerow(["Remaining", f"{report_data.budget_remaining:.2f}"])
                    writer.writerow(["Usage (%)", f"{report_data.budget_percentage:.1f}"])
                    writer.writerow([])

                # Категории
                if options.include_categories_summary and report_data.spending_categories:
                    writer.writerow(["SPENDING BY CATEGORIES"])
                    writer.writerow(["Category", "Amount", "Percentage", "Transactions"])
                    for category in report_data.spending_categories:
                        writer.writerow([
                            category.category_name,
                            f"{category.amount:.2f}",
                            f"{category.percentage:.1f}",
                            category.transaction_count
                        ])
                    writer.writerow([])

                # Данные для графика доходы vs расходы
                if options.include_charts and report_data.income_vs_expenses_chart and report_data.income_vs_expenses_chart.data_points:
                    writer.writerow(["INCOME VS EXPENSES DATA"])
                    writer.writerow(["Date", "Income", "Expenses", "Net"])
                    for point in report_data.income_vs_expenses_chart.data_points:
                        writer.writerow([
                            point.date.strftime('%Y-%m-%d'),
                            f"{point.income:.2f}",
                            f"{point.expenses:.2f}",
                            f"{point.net:.2f}"
                        ])
                    writer.writerow([])

                # Недельный тренд
                if report_data.weekly_trend and report_data.weekly_trend.weeks:
                    writer.writerow(["WEEKLY TREND"])
                    writer.writerow(["Week Start", "Week End", "Expenses"])
                    for week in report_data.weekly_trend.weeks:
                        writer.writerow([
                            week.week_start.strftime('%Y-%m-%d'),
                            week.week_end.strftime('%Y-%m-%d'),
                            f"{week.total_expenses:.2f}"
                        ])
                    writer.writerow(["Average Weekly Spending", "", f"{report_data.weekly_trend.average_weekly_spending:.2f}"])
                    writer.writerow([])

                # Инсайты
                if options.include_insights and report_data.insights:
                    writer.writerow(["FINANCIAL INSIGHTS"])
                    writer.writerow(["Title", "Message", "Type"])
                    for insight in report_data.insights:
                        writer.writerow([insight.title, insight.message, insight.type])

            logger.info(f"CSV report created successfully")
            return file_path

        except Exception as e:
            logger.error(f"Error creating CSV report: {e}")
            raise


class ExcelReportGenerator:
    @staticmethod
    async def generate(
            report_data: FinancialReportSummary,
            options: ExportOptions,
            export_dir: str,
            export_id: str
    ) -> str:
        """Генерация Excel отчета"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            from openpyxl.chart import PieChart, LineChart, Reference
        except ImportError:
            logger.warning("openpyxl not available, falling back to CSV")
            return await CSVReportGenerator.generate(report_data, options, export_dir, export_id)

        file_path = os.path.join(export_dir, f"financial_report_{export_id}.xlsx")
        logger.info(f"Generating Excel report: {file_path}")

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Financial Report"

            # Стили
            header_font = Font(bold=True, size=14)
            title_font = Font(bold=True, size=16)

            row = 1

            # Заголовок
            ws.cell(row=row, column=1, value=f"Financial Report - {report_data.period_name}").font = title_font
            row += 2

            # Основные показатели
            ws.cell(row=row, column=1, value="Financial Summary").font = header_font
            row += 1

            summary_data = [
                ("Income", report_data.income),
                ("Expenses", report_data.expenses),
                ("Net Balance", report_data.net_balance),
                ("Savings", report_data.savings),
                ("Savings Rate (%)", report_data.savings_rate),
                ("Transaction Count", report_data.transaction_count),
            ]

            for metric, value in summary_data:
                ws.cell(row=row, column=1, value=metric)
                ws.cell(row=row, column=2, value=value)
                row += 1

            row += 2

            # Категории расходов
            if options.include_categories_summary and report_data.spending_categories:
                ws.cell(row=row, column=1, value="Spending by Categories").font = header_font
                row += 1

                # Заголовки
                headers = ["Category", "Amount", "Percentage", "Transactions"]
                for col, header in enumerate(headers, 1):
                    ws.cell(row=row, column=col, value=header).font = Font(bold=True)
                row += 1

                # Данные
                for category in report_data.spending_categories:
                    ws.cell(row=row, column=1, value=category.category_name)
                    ws.cell(row=row, column=2, value=category.amount)
                    ws.cell(row=row, column=3, value=category.percentage)
                    ws.cell(row=row, column=4, value=category.transaction_count)
                    row += 1

            # Сохраняем файл
            wb.save(file_path)
            logger.info(f"Excel report created successfully")
            return file_path

        except Exception as e:
            logger.error(f"Error creating Excel report: {e}")
            raise